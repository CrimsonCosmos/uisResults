#!/usr/bin/env python3
"""
UIS Athletics Results Scraper
Scrapes athletic.net for Illinois-Springfield athletes' recent results.
Checks each athlete's profile for events in the last 5 days and identifies PRs, SRs.

This scraper uses a hybrid approach:
1. First tries fast API calls (10-50x faster)
2. Falls back to Selenium scraping if API fails
"""

import time
import re
import json
import requests
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd


# Events that are comparable between indoor and outdoor track
# Indoor SR should carry over to outdoor for these events
COMPARABLE_INDOOR_OUTDOOR_EVENTS = {
    '800', '800 Meters', '800m',
    '1000', '1000 Meters', '1000m',
    '1500', '1500 Meters', '1500m',
    'Mile', '1 Mile',
    '3000', '3000 Meters', '3000m',
    '5000', '5000 Meters', '5000m', '5K',
    '10000', '10000 Meters', '10000m', '10K',
    'High Jump', 'HJ',
    'Pole Vault', 'PV',
    'Long Jump', 'LJ',
    'Triple Jump', 'TJ',
    'Shot Put', 'SP',
}

# NCAA D2 Qualifying Standards (2025-26 Season)
# Source: https://www.ncaa.org/sports/2013/11/5/division-ii-men-s-and-women-s-indoor-track-and-field.aspx
# Times are in seconds, distances in meters
NCAA_D2_STANDARDS = {
    'indoor': {
        'M': {  # Men's Indoor
            '60': 6.85, '60 Meters': 6.85, '60m': 6.85,
            '200': 21.65, '200 Meters': 21.65, '200m': 21.65,
            '400': 48.30, '400 Meters': 48.30, '400m': 48.30,
            '800': 112.01, '800 Meters': 112.01, '800m': 112.01,  # 1:52.01
            'Mile': 247.63, '1 Mile': 247.63,  # 4:07.63
            '3000': 492.89, '3000 Meters': 492.89, '3000m': 492.89,  # 8:12.89
            '5000': 861.50, '5000 Meters': 861.50, '5000m': 861.50, '5K': 861.50,  # 14:21.50
            '60 Hurdles': 8.05, '60H': 8.05, '60m Hurdles': 8.05,
            'High Jump': 2.08, 'HJ': 2.08,
            'Pole Vault': 4.95, 'PV': 4.95,
            'Long Jump': 7.35, 'LJ': 7.35,
            'Triple Jump': 14.80, 'TJ': 14.80,
            'Shot Put': 17.20, 'SP': 17.20,
            'Weight Throw': 19.00, 'WT': 19.00,
        },
        'W': {  # Women's Indoor
            '60': 7.61, '60 Meters': 7.61, '60m': 7.61,
            '200': 24.63, '200 Meters': 24.63, '200m': 24.63,
            '400': 56.50, '400 Meters': 56.50, '400m': 56.50,
            '800': 133.30, '800 Meters': 133.30, '800m': 133.30,  # 2:13.30
            'Mile': 294.48, '1 Mile': 294.48,  # 4:54.48
            '3000': 590.34, '3000 Meters': 590.34, '3000m': 590.34,  # 9:50.34
            '5000': 1023.21, '5000 Meters': 1023.21, '5000m': 1023.21, '5K': 1023.21,  # 17:03.21
            '60 Hurdles': 8.79, '60H': 8.79, '60m Hurdles': 8.79,
            'High Jump': 1.67, 'HJ': 1.67,
            'Pole Vault': 3.72, 'PV': 3.72,
            'Long Jump': 5.72, 'LJ': 5.72,
            'Triple Jump': 11.74, 'TJ': 11.74,
            'Shot Put': 13.64, 'SP': 13.64,
            'Weight Throw': 17.26, 'WT': 17.26,
        },
    },
    'outdoor': {
        'M': {  # Men's Outdoor
            '100': 10.55, '100 Meters': 10.55, '100m': 10.55,
            '200': 21.30, '200 Meters': 21.30, '200m': 21.30,
            '400': 47.20, '400 Meters': 47.20, '400m': 47.20,
            '800': 110.50, '800 Meters': 110.50, '800m': 110.50,  # 1:50.50
            '1500': 228.00, '1500 Meters': 228.00, '1500m': 228.00,  # 3:48.00
            '5000': 855.00, '5000 Meters': 855.00, '5000m': 855.00, '5K': 855.00,  # 14:15.00
            '10000': 1770.00, '10000 Meters': 1770.00, '10000m': 1770.00, '10K': 1770.00,  # 29:30.00
            '110 Hurdles': 14.30, '110H': 14.30, '110m Hurdles': 14.30,
            '400 Hurdles': 52.50, '400H': 52.50, '400m Hurdles': 52.50,
            'Steeplechase': 555.00, '3000 Steeplechase': 555.00, '3000m Steeplechase': 555.00,  # 9:15.00
            'High Jump': 2.10, 'HJ': 2.10,
            'Pole Vault': 5.00, 'PV': 5.00,
            'Long Jump': 7.45, 'LJ': 7.45,
            'Triple Jump': 15.00, 'TJ': 15.00,
            'Shot Put': 17.50, 'SP': 17.50,
            'Discus': 52.00, 'Discus Throw': 52.00,
            'Hammer': 56.00, 'Hammer Throw': 56.00,
            'Javelin': 62.00, 'Javelin Throw': 62.00,
            'Decathlon': 7000,
        },
        'W': {  # Women's Outdoor
            '100': 11.75, '100 Meters': 11.75, '100m': 11.75,
            '200': 24.20, '200 Meters': 24.20, '200m': 24.20,
            '400': 55.50, '400 Meters': 55.50, '400m': 55.50,
            '800': 130.00, '800 Meters': 130.00, '800m': 130.00,  # 2:10.00
            '1500': 270.00, '1500 Meters': 270.00, '1500m': 270.00,  # 4:30.00
            '5000': 1005.00, '5000 Meters': 1005.00, '5000m': 1005.00, '5K': 1005.00,  # 16:45.00
            '10000': 2100.00, '10000 Meters': 2100.00, '10000m': 2100.00, '10K': 2100.00,  # 35:00.00
            '100 Hurdles': 14.00, '100H': 14.00, '100m Hurdles': 14.00,
            '400 Hurdles': 61.00, '400H': 61.00, '400m Hurdles': 61.00,
            'Steeplechase': 660.00, '3000 Steeplechase': 660.00, '3000m Steeplechase': 660.00,  # 11:00.00
            'High Jump': 1.70, 'HJ': 1.70,
            'Pole Vault': 3.80, 'PV': 3.80,
            'Long Jump': 5.85, 'LJ': 5.85,
            'Triple Jump': 12.00, 'TJ': 12.00,
            'Shot Put': 14.00, 'SP': 14.00,
            'Discus': 47.00, 'Discus Throw': 47.00,
            'Hammer': 54.00, 'Hammer Throw': 54.00,
            'Javelin': 44.00, 'Javelin Throw': 44.00,
            'Heptathlon': 4800,
        },
    },
}


def get_ncaa_standard(event_name, sport, gender):
    """
    Get NCAA D2 qualifying standard for an event.
    Returns the standard in seconds (for time events) or meters (for field events).
    Returns None if no standard found.
    """
    season = 'indoor' if sport == 'indoor' else 'outdoor'
    standards = NCAA_D2_STANDARDS.get(season, {}).get(gender, {})

    # Try exact match first
    if event_name in standards:
        return standards[event_name]

    # Try partial match
    event_lower = event_name.lower()
    for std_event, std_value in standards.items():
        if std_event.lower() in event_lower or event_lower in std_event.lower():
            return std_value

    return None


def format_standard_time(seconds):
    """Convert seconds to time string (e.g., 112.01 -> '1:52.01')."""
    if seconds >= 3600:
        hours = int(seconds // 3600)
        mins = int((seconds % 3600) // 60)
        secs = seconds % 60
        return f"{hours}:{mins:02d}:{secs:05.2f}"
    elif seconds >= 60:
        mins = int(seconds // 60)
        secs = seconds % 60
        return f"{mins}:{secs:05.2f}"
    else:
        return f"{seconds:.2f}"


class AthleticNetAPI:
    """
    Fast API client for athletic.net.
    Captures tokens from browser session and makes direct API calls.
    """

    API_BASE = "https://www.athletic.net/api/v1"

    def __init__(self):
        self.session = requests.Session()
        self.tokens = {}
        self.cookies_set = False

    def init_from_browser(self, driver):
        """
        Capture API tokens from browser's network requests.
        Call this after the page has fully loaded.
        """
        # Transfer cookies
        for cookie in driver.get_cookies():
            self.session.cookies.set(cookie['name'], cookie['value'])
        self.cookies_set = True

        # Capture tokens from network logs
        try:
            logs = driver.get_log('performance')
            for log in logs:
                try:
                    message = json.loads(log['message'])['message']
                    if message['method'] == 'Network.requestWillBeSent':
                        headers = message['params']['request'].get('headers', {})
                        if headers.get('anettokens'):
                            self.tokens['anettokens'] = headers['anettokens']
                        if headers.get('anet-site-roles-token'):
                            self.tokens['anet-site-roles-token'] = headers['anet-site-roles-token']
                        if headers.get('anet-appinfo'):
                            self.tokens['anet-appinfo'] = headers['anet-appinfo']
                        if self.tokens.get('anettokens') and self.tokens.get('anet-site-roles-token'):
                            break
                except:
                    pass
        except Exception as e:
            print(f"  Warning: Could not capture API tokens: {e}")

        return bool(self.tokens.get('anettokens'))

    def _make_request(self, endpoint, params=None, referer=None):
        """Make API request with captured tokens."""
        if not self.cookies_set:
            return None

        url = f"{self.API_BASE}/{endpoint}"
        headers = {
            'Accept': 'application/json, text/plain, */*',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
            'anet-appinfo': self.tokens.get('anet-appinfo', 'web:web:0:360'),
        }

        if self.tokens.get('anettokens'):
            headers['anettokens'] = self.tokens['anettokens']
        if self.tokens.get('anet-site-roles-token'):
            headers['anet-site-roles-token'] = self.tokens['anet-site-roles-token']
        if referer:
            headers['Referer'] = referer

        try:
            resp = self.session.get(url, params=params, headers=headers, timeout=10)
            if resp.status_code == 200:
                return resp.json()
        except Exception as e:
            pass
        return None

    def get_roster(self, season_id, referer=None):
        """Get team roster via API."""
        data = self._make_request(
            "TeamHome/GetAthletes",
            params={'seasonId': season_id},
            referer=referer
        )
        if data and isinstance(data, list):
            return [{'id': str(a['ID']), 'name': a['Name'], 'gender': a.get('Gender', '')} for a in data]
        return None

    def get_athlete_bio(self, athlete_id, sport='xc', referer=None):
        """Get athlete bio data including all results."""
        sport_code = 'xc' if sport == 'xc' else 'tf'
        data = self._make_request(
            "AthleteBio/GetAthleteBioData",
            params={'athleteId': athlete_id, 'sport': sport_code, 'level': 0},
            referer=referer
        )
        return data

    def parse_athlete_results(self, bio_data, athlete_id, athlete_name, cutoff_date, year):
        """
        Parse API athlete bio response into results and bests format.
        Returns (results, bests) tuple matching the Selenium scraper format.
        """
        results = []
        bests = {}

        if not bio_data:
            return results, bests

        # Parse results from bio data
        # API returns results with PersonalBest and SeasonBest flags
        results_data = bio_data.get('results', [])

        for result in results_data:
            try:
                # Parse date
                date_str = result.get('MeetDate', '')
                if date_str:
                    # API returns dates in format "2025-01-15T00:00:00"
                    result_date = datetime.strptime(date_str[:10], "%Y-%m-%d")
                else:
                    continue

                # Skip if before cutoff
                if result_date < cutoff_date:
                    continue

                event = result.get('Event', '')
                time_str = result.get('Result', '')
                place = result.get('Place', 0)
                meet_name = result.get('MeetName', '')

                # Determine record type
                record_type = None
                if result.get('PersonalBest'):
                    record_type = 'PR'
                elif result.get('SeasonBest'):
                    record_type = 'SR'

                results.append({
                    'athlete_name': athlete_name,
                    'athlete_id': athlete_id,
                    'event': event,
                    'place': place,
                    'time': time_str,
                    'record_type': record_type,
                    'date': result_date,
                    'date_str': result_date.strftime("%b %d, %Y"),
                    'meet_name': meet_name
                })

            except Exception:
                continue

        # Parse bests from bio data
        # Look for season bests and personal bests per event
        season_bests = bio_data.get('seasonBests', [])
        personal_bests = bio_data.get('personalBests', [])

        # Build bests dictionary from personal bests
        for pb in personal_bests:
            event = pb.get('Event', '')
            time_str = pb.get('Result', '')
            if event and time_str:
                time_secs = self._time_to_seconds(time_str)
                bests[event] = {
                    'pr': time_str,
                    'pr_seconds': time_secs
                }

        # Add season bests
        for sb in season_bests:
            event = sb.get('Event', '')
            time_str = sb.get('Result', '')
            sb_year = sb.get('Year', 0)

            if event and time_str and sb_year == year:
                time_secs = self._time_to_seconds(time_str)
                if event not in bests:
                    bests[event] = {}
                bests[event]['sr'] = time_str
                bests[event]['sr_seconds'] = time_secs

        return results, bests

    def _time_to_seconds(self, time_str):
        """Convert time string to seconds."""
        if not time_str:
            return float('inf')

        time_str = re.sub(r'[PRSRprsr\s\*]+', '', time_str).strip()

        try:
            if ':' in time_str:
                parts = time_str.split(':')
                if len(parts) == 2:
                    return int(parts[0]) * 60 + float(parts[1])
                elif len(parts) == 3:
                    return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
            else:
                return float(time_str)
        except (ValueError, IndexError):
            return float('inf')

    # ===== NEW MEET-BASED APPROACH (MUCH FASTER) =====

    @staticmethod
    def get_season_id(sport, year):
        """
        Convert sport + year to athletic.net's seasonId format.

        athletic.net uses different seasonId formats:
        - Cross Country: year (e.g., 2025)
        - Indoor Track: year + 10000 (e.g., 12026 for 2026)
        - Outdoor Track: year (e.g., 2025)
        """
        if sport == 'indoor':
            return year + 10000
        else:
            return year

    def get_team_calendar(self, season_id, referer=None):
        """
        Get team calendar with all meets for a season.
        Returns list of meets with IDs, dates, and hasResults flags.
        """
        data = self._make_request(
            "TeamHomeCal/GetCalendar",
            params={'seasonId': season_id},
            referer=referer
        )
        if data and isinstance(data, list):
            return data
        return None

    def get_meet_data(self, meet_id, sport='xc', referer=None):
        """
        Get meet data including division IDs.
        Returns dict with xcDivisions or tfDivisions containing division info.
        """
        sport_code = 'xc' if sport == 'xc' else 'tf'
        data = self._make_request(
            "Meet/GetMeetData",
            params={'meetId': meet_id, 'sport': sport_code},
            referer=referer
        )
        return data

    def get_meet_results(self, div_id, meet_id, referer=None):
        """
        Get all results for a meet division via POST to GetResultsData3.
        This is the FAST way to get results - one call per division instead of per athlete!

        Returns dict with 'resultsXC' or 'resultsTF' list containing all athlete results
        with isPr and isSr flags.
        """
        if not self.cookies_set:
            return None

        url = f"{self.API_BASE}/Meet/GetResultsData3"
        headers = {
            'Accept': 'application/json, text/plain, */*',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
            'Content-Type': 'application/json',
            'anet-appinfo': self.tokens.get('anet-appinfo', 'web:web:0:360'),
        }

        if self.tokens.get('anettokens'):
            headers['anettokens'] = self.tokens['anettokens']
        if self.tokens.get('anet-site-roles-token'):
            headers['anet-site-roles-token'] = self.tokens['anet-site-roles-token']
        if referer:
            headers['Referer'] = referer

        try:
            resp = self.session.post(url, json={'divId': div_id}, headers=headers, timeout=15)
            if resp.status_code == 200:
                return resp.json()
        except Exception as e:
            pass
        return None

    def get_team_results_from_meets(self, team_id, season_id, sport, cutoff_date, driver, referer=None):
        """
        FAST approach: Get all team results by checking recent meets.

        Instead of checking each athlete individually, we:
        1. Get the team calendar (list of meets)
        2. Filter to meets within date range that have results
        3. For each meet, load the page to get meet-specific tokens
        4. Get division results and filter to our team

        This is 10-50x faster than checking individual athletes!

        Args:
            driver: Selenium WebDriver instance (needed to load meet pages for tokens)

        Returns list of results in the standard format.
        """
        results = []

        # Get team calendar
        calendar = self.get_team_calendar(season_id, referer=referer)
        if not calendar:
            return None

        # Filter to meets with results within date range
        recent_meets = []
        for meet in calendar:
            if not meet.get('MeetHasResults'):
                continue

            # Parse meet date
            meet_date_str = meet.get('StartDate', '')
            if not meet_date_str:
                continue

            try:
                meet_date = datetime.strptime(meet_date_str[:10], "%Y-%m-%d")
                if meet_date >= cutoff_date:
                    recent_meets.append({
                        'id': meet['MeetID'],
                        'name': meet['Name'],
                        'date': meet_date,
                        'date_str': meet_date.strftime("%b %d, %Y")
                    })
            except:
                continue

        if not recent_meets:
            return results  # Empty list, no recent meets

        print(f"  Found {len(recent_meets)} recent meet(s) with results")

        # For each recent meet, get results
        for meet in recent_meets:
            meet_id = meet['id']
            meet_name = meet['name']
            meet_date = meet['date']
            meet_date_str = meet['date_str']

            print(f"    Checking {meet_name}...", end=' ', flush=True)

            # Build meet URL based on sport
            if sport == 'xc':
                meet_url = f"https://www.athletic.net/CrossCountry/meet/{meet_id}/results"
            else:
                meet_url = f"https://www.athletic.net/TrackAndField/meet/{meet_id}/results"

            # IMPORTANT: Load the meet page to get meet-specific tokens
            # The anettokens JWT contains the meetId and is required for GetResultsData3
            try:
                driver.get(meet_url)
                time.sleep(2)  # Wait for page and API calls

                # Capture fresh tokens for this meet
                logs = driver.get_log('performance')
                for log in logs:
                    try:
                        message = json.loads(log['message'])['message']
                        if message['method'] == 'Network.requestWillBeSent':
                            headers = message['params']['request'].get('headers', {})
                            if headers.get('anettokens'):
                                self.tokens['anettokens'] = headers['anettokens']
                            if headers.get('anet-site-roles-token'):
                                self.tokens['anet-site-roles-token'] = headers['anet-site-roles-token']
                    except:
                        pass
            except Exception as e:
                print(f"page load failed: {e}")
                continue

            # Get meet data to find divisions
            meet_data = self.get_meet_data(meet_id, sport=sport, referer=meet_url)
            if not meet_data:
                print("no data")
                continue

            # Get divisions based on sport
            divisions = meet_data.get('xcDivisions', []) if sport == 'xc' else meet_data.get('tfDivisions', [])

            if not divisions:
                print("no divisions")
                continue

            meet_results_count = 0

            # Get results for each division
            for div in divisions:
                div_id = div.get('IDMeetDiv')
                div_name = div.get('DivName', 'Unknown')
                event_name = div_name  # For XC, division name is the event

                if not div_id:
                    continue

                # Get results for this division
                div_results = self.get_meet_results(div_id, meet_id, referer=meet_url)
                if not div_results:
                    continue

                # Extract results (XC uses resultsXC, TF would use resultsTF)
                results_list = div_results.get('resultsXC', []) or div_results.get('resultsTF', [])

                # Filter to our team (handle both int and string comparisons)
                for r in results_list:
                    school_id = r.get('IDSchool')
                    if school_id != team_id and str(school_id) != str(team_id):
                        continue

                    # Build result object
                    first_name = r.get('FirstName', '')
                    last_name = r.get('LastName', '')
                    athlete_name = f"{first_name} {last_name}".strip()

                    record_type = None
                    if r.get('isPr'):
                        record_type = 'PR'
                    elif r.get('isSr'):
                        record_type = 'SR'

                    result_obj = {
                        'athlete_name': athlete_name,
                        'athlete_id': str(r.get('AthleteID', '')),
                        'event': event_name,
                        'event_id': r.get('EventID'),  # For track event matching
                        'place': r.get('Place', 0),
                        'time': r.get('Result', ''),
                        'record_type': record_type,
                        'date': meet_date,
                        'date_str': meet_date_str,
                        'meet_name': meet_name
                    }

                    results.append(result_obj)
                    meet_results_count += 1

            if meet_results_count > 0:
                print(f"found {meet_results_count} UIS result(s)")
            else:
                print("no UIS results")

        # Fetch bests for ALL athletes who competed
        # This gives us previous PRs for PR results and current bests for others
        if results:
            unique_athletes = len(set(r['athlete_id'] for r in results))
            print(f"  Fetching athlete bests for {unique_athletes} athletes...")

            # Reload team page to get fresh tokens (meet tokens might not work for bio requests)
            sport_path = 'cross-country' if sport == 'xc' else 'track-and-field'
            team_url = f"https://www.athletic.net/team/65580/{sport_path}/{datetime.now().year}"
            driver.get(team_url)
            time.sleep(2)
            logs = driver.get_log('performance')
            for log in logs:
                try:
                    message = json.loads(log['message'])['message']
                    if message['method'] == 'Network.requestWillBeSent':
                        headers = message['params']['request'].get('headers', {})
                        if headers.get('anettokens'):
                            self.tokens['anettokens'] = headers['anettokens']
                        if headers.get('anet-site-roles-token'):
                            self.tokens['anet-site-roles-token'] = headers['anet-site-roles-token']
                except:
                    pass

            self._fetch_athlete_bests(results, sport, driver)

        return results

    def get_track_results_from_athletes(self, roster, season_id, sport, cutoff_date, driver, referer=None):
        """
        Get track results by checking each athlete's bio.
        This is needed because track meet results API doesn't return data the same way as XC.

        Returns list of results with PR/SR flags.
        """
        results = []
        cutoff_str = cutoff_date.strftime('%Y-%m-%d')

        print(f"  Checking roster ({len(roster)} total athletes)...")

        active_count = 0  # Count of athletes with current season activity
        for i, athlete in enumerate(roster):
            athlete_id = athlete['id']
            athlete_name = athlete['name']
            athlete_gender = athlete.get('gender', '')  # 'M' or 'F'

            # Get athlete bio with retry and rate limiting
            bio_data = None
            for attempt in range(7):  # Up to 7 attempts
                bio_data = self.get_athlete_bio(athlete_id, sport=sport, referer=referer)
                if bio_data:
                    break
                wait_time = 1.0 + (attempt * 0.5)  # Increasing backoff: 1s, 1.5s, 2s, etc.
                time.sleep(wait_time)

            if not bio_data:
                print(f"      [!] Failed to get data for {athlete_name} after {attempt+1} attempts")
                continue

            # Get results and event mappings
            tf_results = bio_data.get('resultsTF', [])

            # Quick check: does this athlete have ANY results this season?
            has_current_season = any(r.get('SeasonID') == season_id for r in tf_results)
            if not has_current_season:
                # Skip athletes not on current season roster (no API delay needed)
                continue

            active_count += 1
            # Progress indicator - only show athletes with current season activity
            print(f"    [{active_count}] {athlete_name}")

            # Delay between athletes to avoid rate limiting
            time.sleep(0.25)

            events_list = bio_data.get('eventsTF', [])
            meets_data = bio_data.get('meets', {})

            # Build lookup maps
            events = {}  # EventID -> event name
            event_name_to_ids = {}  # event name -> list of EventIDs (for matching across indoor/outdoor)
            for e in events_list:
                if isinstance(e, dict):
                    event_id = e.get('IDEvent')
                    name = e.get('Event', f"Event {event_id}")
                    events[event_id] = name
                    # Group EventIDs by normalized event name (for FT detection across seasons)
                    if name not in event_name_to_ids:
                        event_name_to_ids[name] = []
                    event_name_to_ids[name].append(event_id)

            # meets_data is a dict with string keys (meet IDs)
            # Each value is a dict with 'IDMeet', 'MeetName', 'EndDate'
            meets = {}
            if isinstance(meets_data, dict):
                for meet_id, meet_info in meets_data.items():
                    if isinstance(meet_info, dict):
                        meets[int(meet_id)] = meet_info
                        meets[str(meet_id)] = meet_info  # Also store with string key

            # Filter for this season and after cutoff date
            for r in tf_results:
                if r.get('SeasonID') != season_id:
                    continue

                result_date = r.get('ResultDate', '')[:10] if r.get('ResultDate') else ''
                if result_date < cutoff_str:
                    continue

                # Get event and meet info
                event_name = events.get(r.get('EventID'), f"Event {r.get('EventID')}")
                meet_id = r.get('MeetID')
                meet_info = meets.get(meet_id) or meets.get(str(meet_id)) or {}
                meet_name = meet_info.get('MeetName', 'Unknown Meet') if isinstance(meet_info, dict) else 'Unknown Meet'

                # Determine record type
                record_type = None
                is_pr = r.get('PersonalBest')
                is_sr = r.get('SeasonBest')
                # PersonalBest can be a number (98 = PR) or boolean
                if is_pr and (is_pr is True or is_pr >= 90):
                    record_type = 'PR'
                elif is_sr and (is_sr is True or is_sr >= 1):
                    record_type = 'SR'

                # Find previous best for this event from athlete's history
                current_result = r.get('Result', '')
                event_id = r.get('EventID')
                previous_pr = None
                previous_sr = None

                # Helper function to parse time strings
                def parse_time(time_str):
                    """Convert time string to seconds for comparison."""
                    if not time_str:
                        return float('inf')
                    # Remove suffixes like 'a', 'h', etc.
                    time_str = re.sub(r'[a-zA-Z\s\*]+', '', str(time_str)).strip()
                    try:
                        if ':' in time_str:
                            parts = time_str.split(':')
                            if len(parts) == 2:
                                return float(parts[0]) * 60 + float(parts[1])
                            elif len(parts) == 3:
                                return float(parts[0]) * 3600 + float(parts[1]) * 60 + float(parts[2])
                        return float(time_str)
                    except:
                        return float('inf')

                # Get all results for this event (match by event NAME to include indoor/outdoor)
                # This ensures FT detection works across seasons (indoor 800 + outdoor 800)
                matching_event_ids = event_name_to_ids.get(event_name, [event_id])
                event_results = [
                    res for res in tf_results
                    if res.get('EventID') in matching_event_ids and res.get('Result') != current_result
                ]

                if event_results:
                    # Find previous PR (best time before this result)
                    all_times = [(parse_time(res.get('Result')), res.get('Result')) for res in event_results]
                    all_times = [(t, r) for t, r in all_times if t != float('inf')]
                    if all_times:
                        all_times.sort(key=lambda x: x[0])
                        previous_pr = all_times[0][1]  # Best previous time

                    # Find previous SR (best time this season before this result)
                    # For outdoor season, also include indoor results for comparable events
                    valid_season_ids = {season_id}

                    # Check if this is outdoor season and event is comparable
                    # Outdoor season_id is just the year (e.g., 2026)
                    # Indoor season_id is year + 10000 (e.g., 12026)
                    is_outdoor = season_id < 10000
                    is_comparable_event = any(comp.lower() in event_name.lower()
                                              for comp in COMPARABLE_INDOOR_OUTDOOR_EVENTS)

                    if is_outdoor and is_comparable_event:
                        # Include indoor season from same academic year
                        indoor_season_id = season_id + 10000
                        valid_season_ids.add(indoor_season_id)

                    season_results = [
                        res for res in event_results
                        if res.get('SeasonID') in valid_season_ids
                    ]
                    if season_results:
                        season_times = [(parse_time(res.get('Result')), res.get('Result')) for res in season_results]
                        season_times = [(t, r) for t, r in season_times if t != float('inf')]
                        if season_times:
                            season_times.sort(key=lambda x: x[0])
                            previous_sr = season_times[0][1]

                # Calculate improvement percentages
                pr_improvement = 0
                sr_improvement = 0
                current_time = parse_time(current_result)

                if previous_pr and current_time != float('inf'):
                    prev_pr_time = parse_time(previous_pr)
                    if prev_pr_time != float('inf') and prev_pr_time > 0:
                        pr_improvement = ((prev_pr_time - current_time) / prev_pr_time) * 100

                if previous_sr and current_time != float('inf'):
                    prev_sr_time = parse_time(previous_sr)
                    if prev_sr_time != float('inf') and prev_sr_time > 0:
                        sr_improvement = ((prev_sr_time - current_time) / prev_sr_time) * 100

                # If it's marked as PR but there's no previous result, it's a First Time (FT)
                if record_type == 'PR' and not previous_pr:
                    record_type = 'FT'
                # If it's marked as PR but the time is worse than previous, it's not actually a PR
                elif record_type == 'PR' and pr_improvement < 0:
                    record_type = None

                # For "Other Results", calculate distance from PR (how much slower)
                distance_from_pr = None
                if record_type is None and previous_pr and pr_improvement < 0:
                    distance_from_pr = abs(pr_improvement)  # Positive value = % slower than PR

                # Calculate NCAA D2 qualifying standard difference
                # Normalize gender: API returns 'M'/'F', some may be empty
                gender = 'M' if athlete_gender in ('M', 'Male') else 'W' if athlete_gender in ('F', 'Female', 'W') else ''
                ncaa_standard = get_ncaa_standard(event_name, sport, gender) if gender else None
                ncaa_diff = None
                ncaa_diff_pct = None

                if ncaa_standard and current_time != float('inf'):
                    # For time events, negative diff means faster than standard (good)
                    # For field events, we'd need to handle differently (higher/longer is better)
                    is_field_event = any(f in event_name.lower() for f in ['jump', 'vault', 'put', 'throw', 'discus', 'hammer', 'javelin'])

                    if is_field_event:
                        # Field events: result is in meters, higher is better
                        # Parse the result as a distance
                        try:
                            result_distance = float(re.sub(r'[^\d.]', '', current_result.replace('m', '')))
                            ncaa_diff = result_distance - ncaa_standard  # Positive = over standard
                            if ncaa_standard > 0:
                                ncaa_diff_pct = (ncaa_diff / ncaa_standard) * 100
                        except:
                            pass
                    else:
                        # Time events: lower is better
                        ncaa_diff = current_time - ncaa_standard  # Negative = under standard (qualified!)
                        if ncaa_standard > 0:
                            ncaa_diff_pct = (ncaa_diff / ncaa_standard) * 100

                results.append({
                    'athlete_id': athlete_id,
                    'athlete_name': athlete_name,
                    'event': event_name,
                    'time': r.get('Result', ''),
                    'place': r.get('Place', ''),
                    'date_str': result_date,
                    'meet_name': meet_name,
                    'record_type': record_type,
                    'previous_pr': previous_pr,
                    'previous_sr': previous_sr,
                    'pr_improvement': pr_improvement,
                    'sr_improvement': sr_improvement,
                    'distance_from_pr': distance_from_pr,
                    'gender': gender,
                    'ncaa_standard': ncaa_standard,
                    'ncaa_diff': ncaa_diff,
                    'ncaa_diff_pct': ncaa_diff_pct,
                })

        print(f"  Found {len(results)} results from {active_count} active athletes")
        return results

    def _process_bests(self, r, times, current_seconds):
        """
        Process historical times to determine previous PRs/SRs.
        Updates the result dict in place.
        """
        if not times:
            if r.get('record_type') == 'PR':
                r['first_at_distance'] = True
            return

        pr_time = times[0]  # Best time ever

        if r.get('record_type') == 'PR':
            if len(times) >= 2:
                prev_pr = times[1]
                r['previous_pr'] = prev_pr['time']
                if prev_pr['seconds'] > 0 and prev_pr['seconds'] != float('inf'):
                    improvement = (prev_pr['seconds'] - current_seconds) / prev_pr['seconds'] * 100
                    r['pr_improvement'] = improvement
            else:
                r['first_at_distance'] = True

        elif r.get('record_type') == 'SR':
            if len(times) >= 2:
                prev_sr = times[1]
                r['previous_sr'] = prev_sr['time']
                if prev_sr['seconds'] > 0 and prev_sr['seconds'] != float('inf'):
                    improvement = (prev_sr['seconds'] - current_seconds) / prev_sr['seconds'] * 100
                    r['sr_improvement'] = improvement
            else:
                r['first_at_distance'] = True

        else:
            r['current_pr'] = pr_time['time']
            r['current_pr_seconds'] = pr_time['seconds']
            if pr_time['seconds'] > 0 and pr_time['seconds'] != float('inf'):
                distance_from_pr = (current_seconds - pr_time['seconds']) / pr_time['seconds'] * 100
                r['distance_from_pr'] = distance_from_pr

    def _fetch_athlete_bests(self, results, sport, driver):
        """
        Fetch best times for ALL athletes who competed.
        Updates results in place with:
        - For PRs: previous_pr (2nd best all-time) or marks as first at distance
        - For SRs: previous_sr (2nd best this season)
        - For Others: current_pr and current_sr for comparison
        """
        # Group results by athlete to minimize API calls
        athletes_to_fetch = {}
        for r in results:
            athlete_id = r['athlete_id']
            if athlete_id not in athletes_to_fetch:
                athletes_to_fetch[athlete_id] = {
                    'name': r['athlete_name'],
                    'results': []
                }
            athletes_to_fetch[athlete_id]['results'].append(r)

        # Fetch bio data for each athlete
        # Note: We reuse the existing tokens from team/meet pages - they work for athlete bios
        for athlete_id, data in athletes_to_fetch.items():
            try:
                # Build athlete URL for referer header
                if sport == 'xc':
                    athlete_url = f"https://www.athletic.net/athlete/{athlete_id}/cross-country"
                else:
                    athlete_url = f"https://www.athletic.net/athlete/{athlete_id}/track-and-field"

                # Get athlete bio using existing tokens (no need to load page)
                bio_data = self.get_athlete_bio(athlete_id, sport=sport, referer=athlete_url)
                if not bio_data:
                    continue

                # Extract results from bio - XC uses resultsXC, TF uses resultsTF
                all_bio_results = bio_data.get('resultsXC', []) if sport == 'xc' else bio_data.get('resultsTF', [])
                if not all_bio_results:
                    continue

                # Group by distance (for XC) or EventID (for track)
                distance_results = {}  # Keyed by distance (XC) or EventID (track)
                for br in all_bio_results:
                    if sport == 'xc':
                        key = br.get('Distance', 0)
                    else:
                        key = br.get('EventID', 0)  # Use EventID for track
                    if key not in distance_results:
                        distance_results[key] = []
                    distance_results[key].append({
                        'time': br.get('Result', ''),
                        'seconds': br.get('SortValue', float('inf')),
                        'is_pr': br.get('PersonalBest', False),
                        'is_sr': br.get('SeasonBest', False),
                        'season': br.get('SeasonID', 0)
                    })

                # For each result, find bests
                for r in data['results']:
                    event = r['event']
                    current_time = r['time']
                    current_seconds = self._time_to_seconds(current_time)

                    # For track, use EventID directly if available
                    if sport != 'xc' and r.get('event_id'):
                        event_id = r['event_id']
                        if event_id in distance_results:
                            times = sorted(distance_results[event_id], key=lambda x: x['seconds'])
                            self._process_bests(r, times, current_seconds)
                        else:
                            # No history for this event - first time
                            if r['record_type'] == 'PR':
                                r['first_at_distance'] = True
                        continue

                    # For XC or when EventID not available: Extract distance from event name
                    target_distance = None
                    event_lower = event.lower()

                    # Try meters first (e.g., "8,000 Meters" -> 8000)
                    distance_match = re.search(r'(\d+,?\d*)\s*(?:meters?|m)', event_lower)
                    if distance_match:
                        target_distance = int(distance_match.group(1).replace(',', ''))

                    # Try miles (e.g., "3 Miles" -> ~4828 meters)
                    if not target_distance:
                        miles_match = re.search(r'(\d+(?:\.\d+)?)\s*miles?', event_lower)
                        if miles_match:
                            miles = float(miles_match.group(1))
                            target_distance = int(miles * 1609.34)

                    # Try kilometer (e.g., "5K" -> 5000)
                    if not target_distance:
                        km_match = re.search(r'(\d+)\s*k\b', event_lower)
                        if km_match:
                            target_distance = int(km_match.group(1)) * 1000

                    if not target_distance:
                        continue

                    # Find matching distance results
                    if target_distance not in distance_results:
                        # Try close matches - use 5% tolerance for conversions
                        tolerance = max(100, target_distance * 0.05)
                        best_match = None
                        best_diff = float('inf')
                        for d in distance_results.keys():
                            diff = abs(d - target_distance)
                            if diff < tolerance and diff < best_diff:
                                best_match = d
                                best_diff = diff
                        if best_match:
                            target_distance = best_match

                    if target_distance not in distance_results:
                        continue

                    # Sort all times for this distance (best first)
                    times = sorted(distance_results[target_distance], key=lambda x: x['seconds'])
                    self._process_bests(r, times, current_seconds)

            except Exception as e:
                # Skip this athlete on error
                continue


class AthleticNetScraper:
    """Scraper for athletic.net team results."""

    TEAM_ID = 65580
    BASE_URL = "https://www.athletic.net"

    # Sport configurations
    SPORTS = {
        'xc': {
            'name': 'Cross Country',
            'url_path': 'cross-country',
            'athlete_path': 'cross-country'
        },
        'indoor': {
            'name': 'Indoor Track & Field',
            'url_path': 'track-and-field-indoor',
            'athlete_path': 'track-and-field-indoor'
        },
        'outdoor': {
            'name': 'Outdoor Track & Field',
            'url_path': 'track-and-field-outdoor',
            'athlete_path': 'track-and-field-outdoor'
        }
    }

    def __init__(self, headless=True, year=2025, sport='xc', days_back=5):
        """Initialize the scraper with Chrome webdriver."""
        self.year = year
        self.sport = sport
        self.days_back = days_back

        if sport not in self.SPORTS:
            raise ValueError(f"Invalid sport: {sport}. Choose from: {', '.join(self.SPORTS.keys())}")

        self.sport_config = self.SPORTS[sport]
        self.team_url = f"{self.BASE_URL}/team/{self.TEAM_ID}/{self.sport_config['url_path']}/{year}"

        self.options = Options()
        if headless:
            self.options.add_argument("--headless")
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--disable-dev-shm-usage")
        self.options.add_argument("--window-size=1920,1080")
        self.options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36")

        self.driver = None
        self.cutoff_date = datetime.now() - timedelta(days=days_back)

    def start_browser(self):
        """Start the Chrome browser."""
        print("Starting browser...")
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=self.options)

    def close_browser(self):
        """Close the browser."""
        if self.driver:
            self.driver.quit()

    def get_roster(self):
        """Get the team roster with athlete IDs."""
        print(f"Fetching roster from: {self.team_url}")
        self.driver.get(self.team_url)

        # Wait for athlete links to appear (max 5 seconds)
        try:
            WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='/athlete/']"))
            )
        except:
            pass  # Continue anyway - page might have no athletes

        soup = BeautifulSoup(self.driver.page_source, 'html.parser')

        athletes = []
        seen_ids = set()

        # Find all athlete links
        athlete_links = soup.find_all('a', href=re.compile(r'/athlete/\d+'))

        for link in athlete_links:
            href = link.get('href', '')
            athlete_id_match = re.search(r'/athlete/(\d+)', href)
            if athlete_id_match:
                athlete_id = athlete_id_match.group(1)
                if athlete_id not in seen_ids:
                    seen_ids.add(athlete_id)
                    name = link.get_text(strip=True)
                    if name:  # Only add if we have a name
                        # Clean up name - remove leading initials stuck to the name
                        # Pattern: "KHKhaniya" -> "Khaniya", "EMElijah" -> "Elijah"
                        cleaned_name = re.sub(r'^[A-Z]{2,3}(?=[A-Z][a-z])', '', name)
                        athletes.append({
                            'id': athlete_id,
                            'name': cleaned_name
                        })

        print(f"Found {len(athletes)} athletes on roster")
        return athletes

    def parse_date(self, date_str):
        """Parse date string like 'Sep 5', 'Sep 27', or 'Apr 17, 2025' into datetime."""
        # Try full date format first (Apr 17, 2025)
        try:
            return datetime.strptime(date_str, "%b %d, %Y")
        except ValueError:
            pass

        try:
            return datetime.strptime(date_str, "%B %d, %Y")
        except ValueError:
            pass

        # Try short format with assumed year
        try:
            date_with_year = f"{date_str}, {self.year}"
            return datetime.strptime(date_with_year, "%b %d, %Y")
        except ValueError:
            pass

        try:
            date_with_year = f"{date_str}, {self.year}"
            return datetime.strptime(date_with_year, "%B %d, %Y")
        except ValueError:
            return None

    def time_to_seconds(self, time_str):
        """Convert time string to seconds for comparison."""
        if not time_str:
            return float('inf')

        # Remove PR/SR markers
        time_str = re.sub(r'[PRSRprsr\s\*]+', '', time_str).strip()

        try:
            # Handle MM:SS.ss format
            if ':' in time_str:
                parts = time_str.split(':')
                if len(parts) == 2:
                    minutes = int(parts[0])
                    seconds = float(parts[1])
                    return minutes * 60 + seconds
                elif len(parts) == 3:
                    hours = int(parts[0])
                    minutes = int(parts[1])
                    seconds = float(parts[2])
                    return hours * 3600 + minutes * 60 + seconds
            else:
                # Handle SS.ss format (sprints)
                return float(time_str)
        except (ValueError, IndexError):
            return float('inf')

        return float('inf')

    def get_athlete_results_and_bests(self, athlete_id, athlete_name):
        """Get an athlete's recent results and their best times from their profile."""
        athlete_url = f"{self.BASE_URL}/athlete/{athlete_id}/{self.sport_config['athlete_path']}"
        self.driver.get(athlete_url)

        # Wait for tables to load (max 3 seconds)
        try:
            WebDriverWait(self.driver, 3).until(
                EC.presence_of_element_located((By.TAG_NAME, "table"))
            )
        except:
            pass  # Continue anyway - might have no results

        soup = BeautifulSoup(self.driver.page_source, 'html.parser')

        return self._parse_athlete_page(soup, athlete_id, athlete_name)

    def get_athletes_parallel(self, athletes, num_tabs=3):
        """
        Check multiple athletes in parallel using browser tabs.
        Returns list of (athlete, results, bests) tuples.
        """
        if not athletes:
            return []

        all_data = []
        original_handle = self.driver.current_window_handle

        # Process in batches
        for batch_start in range(0, len(athletes), num_tabs):
            batch = athletes[batch_start:batch_start + num_tabs]
            handles = []
            athlete_urls = []

            # Open tabs and start loading pages
            for i, athlete in enumerate(batch):
                url = f"{self.BASE_URL}/athlete/{athlete['id']}/{self.sport_config['athlete_path']}"
                athlete_urls.append(url)

                if i == 0:
                    # Use existing tab for first athlete
                    self.driver.get(url)
                    handles.append(self.driver.current_window_handle)
                else:
                    # Open new tab for subsequent athletes
                    self.driver.execute_script("window.open('');")
                    self.driver.switch_to.window(self.driver.window_handles[-1])
                    self.driver.get(url)
                    handles.append(self.driver.current_window_handle)

            # Wait a moment for pages to load
            time.sleep(1.5)

            # Collect results from each tab
            for i, (athlete, handle) in enumerate(zip(batch, handles)):
                self.driver.switch_to.window(handle)

                # Quick wait for content
                try:
                    WebDriverWait(self.driver, 2).until(
                        EC.presence_of_element_located((By.TAG_NAME, "table"))
                    )
                except:
                    pass

                # Check for rate limiting (page shows error or unusual content)
                page_source = self.driver.page_source
                if "rate limit" in page_source.lower() or "too many requests" in page_source.lower():
                    print("\n  [!] Rate limited - waiting 30 seconds...")
                    time.sleep(30)
                    self.driver.get(athlete_urls[i])
                    time.sleep(2)
                    page_source = self.driver.page_source

                soup = BeautifulSoup(page_source, 'html.parser')
                results, bests = self._parse_athlete_page(soup, athlete['id'], athlete['name'])
                all_data.append((athlete, results, bests))

            # Close extra tabs (keep only the first one)
            for handle in handles[1:]:
                self.driver.switch_to.window(handle)
                self.driver.close()

            # Switch back to original tab
            self.driver.switch_to.window(original_handle)

        return all_data

    def _parse_athlete_page(self, soup, athlete_id, athlete_name):
        """Parse an athlete's page HTML and extract results and bests."""

        results = []
        bests = {}  # {event: {'pr': time, 'sr': time, 'pr_seconds': float, 'sr_seconds': float}}

        # Find all tables - results are typically in tables
        tables = soup.find_all('table')

        # We need to find the table with the current season's results
        # It will have dates in format "Sep 5" and times
        for table in tables:
            table_text = table.get_text()

            # Check if this table has recent dates (month names)
            if not re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2}', table_text):
                continue

            # Parse table rows
            rows = table.find_all('tr')
            current_event = None

            for row in rows:
                row_text = row.get_text(separator=' ', strip=True)

                # Check if this row defines an event (e.g., "5000 Meters", "800 Meters")
                event_match = re.match(r'^(\d+(?:,\d+)?\s*(?:Meters?|Mile|Hurdles?|Relay|Steeplechase|Jump|Put|Throw|Vault))', row_text, re.IGNORECASE)
                if event_match:
                    current_event = event_match.group(1).strip()
                    continue

                # Look for result data: place, time, date, meet name
                # Pattern: "1 18:01.1PR Sep 5 Prairie Stars Invitational"
                # Or: "8 16:27.78PR Apr 17, 2025 Bryan Clay Invitational"
                result_pattern = re.compile(
                    r'(\d+)\s+'  # Place
                    r'(\d{1,2}:\d{2}\.\d+|\d+\.\d+)'  # Time
                    r'\s*(PR|SR)?\s*'  # Optional PR/SR marker
                    r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{1,2})(?:,?\s*(\d{4}))?\s+'  # Date with optional year
                    r'(.+?)(?:\s+\d+\s+F)?$',  # Meet name (may end with division info like "1 F")
                    re.IGNORECASE
                )

                match = result_pattern.search(row_text)
                if match and current_event:
                    place = int(match.group(1))
                    time_str = match.group(2)
                    record_type = match.group(3).upper() if match.group(3) else None
                    month = match.group(4)
                    day = match.group(5)
                    year = match.group(6) if match.group(6) else str(self.year)
                    meet_name = match.group(7).strip()

                    # Parse the date
                    date_str = f"{month} {day}, {year}"
                    result_date = self.parse_date(date_str)

                    if result_date and result_date >= self.cutoff_date:
                        results.append({
                            'athlete_name': athlete_name,
                            'athlete_id': athlete_id,
                            'event': current_event,
                            'place': place,
                            'time': time_str,
                            'record_type': record_type,  # 'PR', 'SR', or None
                            'date': result_date,
                            'date_str': date_str,
                            'meet_name': meet_name
                        })

        # Now extract PR/SR bests from the page
        # Look for summary tables that show season/career bests
        # Format: "5000 Meters 2023 Indoor Fr 18:48.71 2024 Outdoor So 17:13.53 * 2025 Indoor Jr 16:37.46 *"
        for table in tables:
            table_text = table.get_text()

            # Try to extract event name at start of table
            event_match = re.match(r'^\s*(\d+(?:,\d+)?\s*(?:Meters?|Mile|Hurdles?|Relay|Steeplechase|Jump|Put|Throw|Vault))', table_text, re.IGNORECASE)
            if not event_match:
                continue

            event = event_match.group(1).strip()

            # Check if this is a summary table (has year + Indoor/Outdoor patterns)
            if not re.search(r'\d{4}\s+(?:Indoor|Outdoor)', table_text):
                continue

            # Find all times with context (year, sport type)
            # Pattern: "2025 Indoor Jr 16:37.46" or "2024 Outdoor So 17:13.53PR"
            time_entries = re.findall(
                r'(\d{4})\s+(Indoor|Outdoor)\s+\w{2}\s+(\d{1,2}:\d{2}\.\d+|\d+\.\d+)\s*(PR)?',
                table_text
            )

            if time_entries:
                # Collect all times with metadata
                all_times = []
                current_season_times = []
                sport_label = 'Indoor' if self.sport == 'indoor' else 'Outdoor'

                for year_str, sport_type, time_str, is_pr in time_entries:
                    secs = self.time_to_seconds(time_str)
                    if secs != float('inf'):
                        entry = {
                            'year': int(year_str),
                            'sport': sport_type,
                            'time': time_str,
                            'seconds': secs,
                            'is_pr': is_pr == 'PR'
                        }
                        all_times.append(entry)

                        # Check if this is current season
                        if int(year_str) == self.year and sport_type == sport_label:
                            current_season_times.append(entry)

                if all_times:
                    # Sort all times to find PR and previous PR
                    all_times.sort(key=lambda x: x['seconds'])
                    best = all_times[0]

                    bests[event] = {
                        'pr': best['time'],
                        'pr_seconds': best['seconds'],
                        'all_times': [t['time'] for t in all_times]  # Keep all times for reference
                    }

                    # Previous PR is the second-best all-time
                    if len(all_times) > 1:
                        bests[event]['previous_pr'] = all_times[1]['time']
                        bests[event]['previous_pr_seconds'] = all_times[1]['seconds']

                    # Current season record (SR)
                    if current_season_times:
                        current_season_times.sort(key=lambda x: x['seconds'])
                        sr = current_season_times[0]
                        bests[event]['sr'] = sr['time']
                        bests[event]['sr_seconds'] = sr['seconds']

                        # Previous SR is second-best this season
                        if len(current_season_times) > 1:
                            bests[event]['previous_sr'] = current_season_times[1]['time']
                            bests[event]['previous_sr_seconds'] = current_season_times[1]['seconds']

        return results, bests

    def get_athlete_bests(self, athlete_id):
        """Get an athlete's PR and SR for each event."""
        athlete_url = f"{self.BASE_URL}/athlete/{athlete_id}/{self.sport_config['athlete_path']}"
        self.driver.get(athlete_url)
        time.sleep(2)

        soup = BeautifulSoup(self.driver.page_source, 'html.parser')

        bests = {}  # {event: {'pr': time, 'sr': time}}

        # Look at the season summary tables (e.g., "5000 Meters 2022 Fr 20:14.1 2023 So 18:57.6 * ...")
        tables = soup.find_all('table')

        for table in tables:
            table_text = table.get_text()

            # Look for patterns like "5000 Meters" followed by yearly results
            event_match = re.match(r'^(\d+(?:,\d+)?\s*(?:Meters?|Mile|Hurdles?|Relay|Steeplechase|Jump|Put|Throw|Vault))', table_text, re.IGNORECASE)
            if event_match:
                event = event_match.group(1).strip()

                # Find all times in this table
                times = re.findall(r'(\d{1,2}:\d{2}\.\d+|\d+\.\d+)(PR)?', table_text)

                if times:
                    # Get the best time (lowest) as PR
                    all_times = [(t[0], t[1] == 'PR') for t in times]
                    time_values = [(self.time_to_seconds(t[0]), t[0], t[1]) for t in all_times]
                    time_values.sort(key=lambda x: x[0])

                    if time_values:
                        best = time_values[0]
                        bests[event] = {
                            'pr': best[1],
                            'pr_seconds': best[0]
                        }

                        # Try to find SR (current year's best) - look for year pattern
                        current_year_pattern = re.compile(rf'{self.year}\s+\w+\s+(\d{{1,2}}:\d{{2}}\.\d+|\d+\.\d+)')
                        sr_match = current_year_pattern.search(table_text)
                        if sr_match:
                            sr_time = sr_match.group(1)
                            bests[event]['sr'] = sr_time
                            bests[event]['sr_seconds'] = self.time_to_seconds(sr_time)

        return bests

    def calculate_improvement(self, current_time_str, previous_time_str):
        """Calculate percentage improvement (lower is better for running)."""
        current = self.time_to_seconds(current_time_str)
        previous = self.time_to_seconds(previous_time_str)

        if previous == float('inf') or previous == 0:
            return 0

        # Improvement is positive when current < previous (faster)
        improvement = (previous - current) / previous * 100
        return improvement

    def run(self):
        """Main execution method."""
        try:
            self.start_browser()

            # Step 1: Get roster
            roster = self.get_roster()

            if not roster:
                print("No athletes found on roster.")
                return None

            all_results = []

            # Step 2: Visit each athlete's profile
            print(f"\nChecking {len(roster)} athletes for results in the last {self.days_back} days...")
            print(f"Cutoff date: {self.cutoff_date.strftime('%Y-%m-%d')}")

            for i, athlete in enumerate(roster):
                print(f"  [{i+1}/{len(roster)}] {athlete['name']}...", end=' ')

                # Get recent results and bests in one page load
                results, bests = self.get_athlete_results_and_bests(athlete['id'], athlete['name'])

                if results:
                    print(f"found {len(results)} recent result(s)")

                    for result in results:
                        event = result['event']
                        current_time = result['time']

                        # Calculate improvements
                        if event in bests:
                            # For PRs: use previous_pr (second-best all-time) since current PR IS the new time
                            previous_pr = bests[event].get('previous_pr')
                            # For SRs: use previous_sr (second-best this season) since current SR IS the new time
                            previous_sr = bests[event].get('previous_sr')
                            # Current SR for non-PR/SR results
                            sr_best = bests[event].get('sr')

                            # For PRs, calculate improvement vs old PR
                            if result['record_type'] == 'PR' and previous_pr:
                                current_seconds = self.time_to_seconds(current_time)
                                prev_pr_seconds = self.time_to_seconds(previous_pr)
                                if prev_pr_seconds != float('inf') and 0.5 < prev_pr_seconds / current_seconds < 2.0:
                                    result['pr_improvement'] = self.calculate_improvement(current_time, previous_pr)
                                    result['previous_pr'] = previous_pr

                            # For SRs, calculate improvement vs old SR
                            if result['record_type'] == 'SR' and previous_sr:
                                current_seconds = self.time_to_seconds(current_time)
                                prev_sr_seconds = self.time_to_seconds(previous_sr)
                                if prev_sr_seconds != float('inf') and 0.5 < prev_sr_seconds / current_seconds < 2.0:
                                    result['sr_improvement'] = self.calculate_improvement(current_time, previous_sr)
                                    result['previous_sr'] = previous_sr

                            # For non-PR/SR, calculate distance from current SR
                            if not result['record_type'] and sr_best:
                                current_seconds = self.time_to_seconds(current_time)
                                sr_seconds = bests[event].get('sr_seconds', float('inf'))
                                if sr_seconds != float('inf'):
                                    # How close (as %) to SR? Lower is closer
                                    result['sr_distance'] = (current_seconds - sr_seconds) / sr_seconds * 100
                                    result['current_sr'] = sr_best

                        all_results.append(result)
                else:
                    print("no recent results")

            if not all_results:
                print("\nNo results found in the specified time period.")
                return None

            # Step 3: Sort results
            print(f"\nFound {len(all_results)} total results. Sorting...")

            # Separate into PRs, SRs, FTs (First Time), others, and DNS/DNF
            prs = [r for r in all_results if r.get('record_type') == 'PR']
            srs = [r for r in all_results if r.get('record_type') == 'SR']
            fts = [r for r in all_results if r.get('record_type') == 'FT']
            dns_dnf = [r for r in all_results if not r.get('record_type') and r.get('time', '').upper() in ['DNS', 'DNF']]
            others = [r for r in all_results if not r.get('record_type') and r.get('time', '').upper() not in ['DNS', 'DNF']]

            # Sort PRs by improvement (highest improvement first)
            prs.sort(key=lambda x: x.get('pr_improvement', 0), reverse=True)

            # Sort SRs by improvement (highest improvement first)
            srs.sort(key=lambda x: x.get('sr_improvement', 0), reverse=True)

            # Sort FTs alphabetically by name
            fts.sort(key=lambda x: x.get('athlete_name', ''))

            # Sort others alphabetically by name
            others.sort(key=lambda x: x.get('athlete_name', ''))

            # Sort DNS/DNF alphabetically by name
            dns_dnf.sort(key=lambda x: x.get('athlete_name', ''))

            # Combine in order: PRs -> SRs -> FTs -> other results -> DNS/DNF
            sorted_results = prs + srs + fts + others + dns_dnf

            # Step 4: Create spreadsheet
            return self.save_to_spreadsheet(sorted_results)

        except Exception as e:
            print(f"Error during scraping: {e}")
            import traceback
            traceback.print_exc()
            raise
        finally:
            self.close_browser()

    def save_to_spreadsheet(self, results, filename=None):
        """Save results to an Excel spreadsheet."""
        if not results:
            print("No results to save.")
            return None

        # Prepare data for DataFrame
        data = []
        for r in results:
            prev_pr = r.get('previous_pr')
            prev_sr = r.get('previous_sr')
            pr_improvement = r.get('pr_improvement', 0)
            sr_improvement = r.get('sr_improvement', 0)

            row = {
                'Name': r['athlete_name'],
                'Type': r.get('record_type', '-'),
                'Event': r['event'],
                'Time/Mark': r['time'],
                'Place': r['place'],
                'Date': r['date_str'],
                'Meet': r['meet_name'],
            }

            # Previous Best (PR) column
            if r.get('record_type') == 'FT' or r.get('first_at_distance'):
                row['Previous Best'] = '-'
            elif prev_pr:
                row['Previous Best'] = prev_pr
            else:
                row['Previous Best'] = '-'

            # Previous Season Best column
            row['Previous SR'] = prev_sr if prev_sr else '-'

            # % Improvement from PR column
            if r.get('record_type') == 'PR' and prev_pr:
                row['% from PR'] = f"{pr_improvement:.2f}%"
            elif r.get('record_type') == 'FT' or not prev_pr:
                row['% from PR'] = '-'
            elif pr_improvement != 0:
                row['% from PR'] = f"{pr_improvement:.2f}%"
            else:
                row['% from PR'] = '-'

            # % Improvement from SR column
            if r.get('record_type') == 'SR' and prev_sr:
                row['% from SR'] = f"{sr_improvement:.2f}%"
            elif not prev_sr:
                row['% from SR'] = '-'
            elif sr_improvement != 0:
                row['% from SR'] = f"{sr_improvement:.2f}%"
            else:
                row['% from SR'] = '-'

            data.append(row)

        df = pd.DataFrame(data)

        columns = ['Name', 'Type', 'Event', 'Time/Mark', 'Place', 'Date', 'Meet', 'Previous Best', 'Previous SR', '% from PR', '% from SR']
        df = df[columns]

        if filename is None:
            sport_name = self.sport_config['name'].replace(' ', '_').replace('&', 'and')
            today = datetime.now().strftime('%Y%m%d')
            filename = f"results_{sport_name}_{self.year}_{today}.xlsx"

        filepath = f"/Users/dylangehl/uisResults/{filename}"
        df.to_excel(filepath, index=False, sheet_name='Results')

        print(f"\nResults saved to: {filepath}")
        print(f"  PRs: {len([r for r in results if r.get('record_type') == 'PR'])}")
        print(f"  SRs: {len([r for r in results if r.get('record_type') == 'SR'])}")
        print(f"  Others: {len([r for r in results if not r.get('record_type')])}")

        return filepath


def get_relevant_sports(days_back):
    """
    Determine which sport/year combos are relevant based on NCAA D2 schedule.

    NCAA D2 Championship dates (approximate):
    - Cross Country: Early November (season: Aug-Nov)
    - Indoor Track: Early March (season: Dec-Mar)
    - Outdoor Track: Late May (season: Mar-Jun)

    We only check sports where meets could have happened in the last N days.
    """
    now = datetime.now()
    current_year = now.year
    month = now.month
    lookback_start = now - timedelta(days=days_back)

    sports = []

    # Cross Country: Season Aug-Nov
    # Check current year if in season, or last year if lookback reaches into last season
    xc_current_start = datetime(current_year, 8, 1)
    xc_current_end = datetime(current_year, 11, 30)
    xc_last_start = datetime(current_year - 1, 8, 1)
    xc_last_end = datetime(current_year - 1, 11, 30)

    if xc_current_start <= now <= xc_current_end:
        # Currently in XC season
        sports.append(('xc', current_year))
    elif lookback_start <= xc_last_end and now >= xc_last_start:
        # Lookback window overlaps with last year's XC season
        sports.append(('xc', current_year - 1))

    # Indoor Track: Season Dec-Mar (spans calendar years)
    # "Indoor 2026" = Dec 2025 through Mar 2026
    if month <= 6:
        # First half of year: current indoor season is this year
        indoor_year = current_year
        indoor_start = datetime(current_year - 1, 12, 1)
        indoor_end = datetime(current_year, 3, 15)
    else:
        # Second half of year: next indoor season starts in Dec
        indoor_year = current_year + 1
        indoor_start = datetime(current_year, 12, 1)
        indoor_end = datetime(current_year + 1, 3, 15)

    if indoor_start <= now <= indoor_end:
        # Currently in indoor season
        sports.append(('indoor', indoor_year))
    elif month <= 6 and lookback_start <= indoor_end:
        # Lookback might catch end of indoor season
        sports.append(('indoor', indoor_year))

    # Outdoor Track: Season Mar-Jun
    outdoor_start = datetime(current_year, 3, 1)
    outdoor_end = datetime(current_year, 6, 15)
    outdoor_last_end = datetime(current_year - 1, 6, 15)

    if outdoor_start <= now <= outdoor_end:
        # Currently in outdoor season
        sports.append(('outdoor', current_year))
    elif lookback_start <= outdoor_last_end and month <= 2:
        # Very long lookback into last year's outdoor (unlikely but possible)
        sports.append(('outdoor', current_year - 1))

    # Remove duplicates while preserving order
    seen = set()
    unique = []
    for s in sports:
        if s not in seen:
            seen.add(s)
            unique.append(s)

    return unique


def _push_results_to_website(data, cutoff_date, end_date, checked_sports, cloud_mode=False):
    """
    Push results as JSON to the wintern-next website repo.
    This triggers a Vercel rebuild to update winterns.com.

    In cloud_mode, just saves JSON locally (for GitHub Actions to handle).
    """
    import json
    import subprocess
    import os
    from datetime import datetime, timezone

    # Build JSON payload
    sport_abbrevs = {'xc': 'Cross Country', 'indoor': 'Indoor Track', 'outdoor': 'Outdoor Track'}
    payload = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "date_range": {
            "start": cutoff_date.strftime("%Y-%m-%d"),
            "end": end_date.strftime("%Y-%m-%d")
        },
        "sports": [sport_abbrevs.get(s, s) for s in checked_sports],
        "results": data
    }

    # Cloud mode: just save JSON locally for GitHub Actions
    if cloud_mode:
        with open("uis-results.json", 'w') as f:
            json.dump(payload, f, indent=2)
        print(f"\nJSON saved to uis-results.json (cloud mode)")
        return

    # Local mode: push to wintern-next repo
    WINTERN_REPO_PATH = os.path.expanduser("~/wintern-next")
    JSON_PATH = os.path.join(WINTERN_REPO_PATH, "public", "data", "uis-results.json")

    # Clone repo if it doesn't exist
    if not os.path.exists(WINTERN_REPO_PATH):
        print("\nCloning wintern-next repo...")
        subprocess.run(
            ["gh", "repo", "clone", "CrimsonCosmos/wintern-next", WINTERN_REPO_PATH],
            check=True, capture_output=True
        )

    # Pull latest changes
    subprocess.run(
        ["git", "-C", WINTERN_REPO_PATH, "pull", "--rebase"],
        capture_output=True
    )

    # Create data directory if needed
    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)

    # Write JSON
    with open(JSON_PATH, 'w') as f:
        json.dump(payload, f, indent=2)

    # Commit and push
    subprocess.run(
        ["git", "-C", WINTERN_REPO_PATH, "add", "public/data/uis-results.json"],
        check=True, capture_output=True
    )

    result = subprocess.run(
        ["git", "-C", WINTERN_REPO_PATH, "diff", "--cached", "--quiet"],
        capture_output=True
    )

    if result.returncode != 0:  # There are changes to commit
        commit_msg = f"Update UIS results ({cutoff_date.strftime('%b %d')} - {end_date.strftime('%b %d')})"
        subprocess.run(
            ["git", "-C", WINTERN_REPO_PATH, "commit", "-m", commit_msg],
            check=True, capture_output=True
        )
        subprocess.run(
            ["git", "-C", WINTERN_REPO_PATH, "push"],
            check=True, capture_output=True
        )
        print(f"\nResults pushed to winterns.com")
    else:
        print(f"\nNo changes to push to website")


def _save_styled_excel(df, filepath, sorted_results):
    """
    Save DataFrame to Excel with professional styling.

    Features:
    - Proper column widths for readability
    - Styled header row (bold, dark background)
    - Color-coded Type column (PR=gold, SR=silver, FT=light blue)
    - Gradient highlighting for Improvement % column
    - Alternating row colors for readability
    - Borders for structure
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Write data to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Define styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    pr_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
    sr_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Silver
    ft_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue

    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # Column widths (index: width)
    column_widths = {
        1: 18,   # Name
        2: 6,    # Type
        3: 31,   # Sport
        4: 20,   # Event
        5: 11,   # Time/Mark
        6: 6,    # Place
        7: 11,   # Date
        8: 28,   # Meet
        9: 11,   # Previous Best
        10: 11,  # Previous SR
        11: 10,  # % from PR
        12: 10,  # % from SR
        13: 10,  # NCAA Std
        14: 10,  # vs NCAA
    }

    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Style header row
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set header row height
    ws.row_dimensions[1].height = 25

    # Style data rows
    type_col = 2  # Type column index
    pr_improvement_col = 11  # % from PR column index
    sr_improvement_col = 12  # % from SR column index

    for row_idx in range(2, len(df) + 2):
        # Set row height
        ws.row_dimensions[row_idx].height = 22

        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

            # Center alignment for most columns, left for Name and Meet
            if col_idx in [1, 8]:  # Name and Meet columns
                cell.alignment = left_align
            else:
                cell.alignment = center_align

        # Color-code the Type column
        type_cell = ws.cell(row=row_idx, column=type_col)
        type_value = type_cell.value
        if type_value == 'PR':
            type_cell.fill = pr_fill
            type_cell.font = Font(bold=True)
        elif type_value == 'SR':
            type_cell.fill = sr_fill
            type_cell.font = Font(bold=True)
        elif type_value == 'FT':
            type_cell.fill = ft_fill
            type_cell.font = Font(bold=True)

        # Alternating row colors (only for non-highlighted cells)
        if row_idx % 2 == 0:
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Don't override Type column fill
                if col_idx != type_col or (type_value not in ['PR', 'SR', 'FT']):
                    if cell.fill == PatternFill():  # No fill yet
                        cell.fill = alt_row_fill

    # Apply gradient to both improvement columns
    def apply_gradient_to_column(col_idx):
        """Apply green gradient to improvement percentage column."""
        improvement_values = []
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value and '%' in str(value) and value != '-':
                try:
                    num_value = float(str(value).replace('%', '').strip())
                    improvement_values.append((row_idx, num_value))
                except ValueError:
                    pass

        if improvement_values:
            values_only = [v for _, v in improvement_values]
            min_val = min(values_only) if values_only else 0
            max_val = max(values_only) if values_only else 1

            for row_idx, value in improvement_values:
                cell = ws.cell(row=row_idx, column=col_idx)

                if max_val > min_val:
                    normalized = (value - min_val) / (max_val - min_val)
                else:
                    normalized = 0.5

                # Gradient: light green to dark green for positive, light red for negative
                if value >= 0:
                    r = int(200 - normalized * (200 - 34))
                    g = int(230 - normalized * (230 - 139))
                    b = int(200 - normalized * (200 - 34))
                else:
                    # Red tint for negative (slower than PR/SR)
                    r = 255
                    g = int(200 + value * 5)  # Gets redder as more negative
                    b = int(200 + value * 5)
                    g = max(150, min(200, g))
                    b = max(150, min(200, b))

                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

                if value >= 0 and normalized > 0.6:
                    cell.font = Font(bold=True, color="FFFFFF")
                else:
                    cell.font = Font(bold=True, color="1F4E79")

    apply_gradient_to_column(pr_improvement_col)
    apply_gradient_to_column(sr_improvement_col)

    # Highlight qualified athletes in "vs NCAA" column (column 14)
    ncaa_col = 14
    event_col = 4  # Event column to check if field event
    qualified_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    close_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light yellow (within 5%)

    for row_idx in range(2, len(df) + 2):
        cell = ws.cell(row=row_idx, column=ncaa_col)
        value = cell.value
        if value and '%' in str(value) and value != '-':
            try:
                # Check if this is a field event (higher is better)
                event_name = str(ws.cell(row=row_idx, column=event_col).value or '').lower()
                is_field_event = any(f in event_name for f in ['jump', 'vault', 'put', 'throw', 'discus', 'hammer', 'javelin'])

                num_value = float(str(value).replace('%', '').replace('+', '').strip())

                # Determine if qualified based on event type
                # Time events: negative % = faster than standard = QUALIFIED
                # Field events: positive % = further/higher than standard = QUALIFIED
                is_qualified = (is_field_event and num_value >= 0) or (not is_field_event and num_value <= 0)
                is_close = abs(num_value) <= 5

                if is_qualified:
                    # Qualified! Highlight green
                    cell.fill = qualified_fill
                    cell.font = Font(bold=True, color="006400")  # Dark green text
                elif is_close:
                    # Close to qualifying (within 5%) - highlight yellow
                    cell.fill = close_fill
                    cell.font = Font(bold=True)
            except ValueError:
                pass

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Save the workbook
    wb.save(filepath)


def main():
    """Main entry point."""
    import argparse

    parser = argparse.ArgumentParser(description='UIS Athletics Results Tracker')
    parser.add_argument('--days', type=int, default=5,
                        help='Number of days back to check (default: 5)')
    parser.add_argument('--headless', action='store_true',
                        help='Run browser in headless mode (may not work with all sites)')

    # Sport filter options (can combine multiple)
    parser.add_argument('--xc', action='store_true',
                        help='Check Cross Country')
    parser.add_argument('--indoor', action='store_true',
                        help='Check Indoor Track & Field')
    parser.add_argument('--outdoor', action='store_true',
                        help='Check Outdoor Track & Field')
    parser.add_argument('--track', action='store_true',
                        help='Check Track & Field (both indoor and outdoor)')

    parser.add_argument('--desktop', action='store_true',
                        help='Save output to Desktop instead of uisResults folder')
    parser.add_argument('--cloud', action='store_true',
                        help='Cloud mode: output JSON to current directory (for GitHub Actions)')

    args = parser.parse_args()

    # Set output directory
    if args.cloud:
        output_dir = "."  # Current directory for GitHub Actions
    elif args.desktop:
        output_dir = "/Users/dylangehl/Desktop"
    else:
        output_dir = "/Users/dylangehl/uisResults"

    print("=" * 70)
    print("UIS Athletics Results Tracker")
    print(f"Checking for results in the last {args.days} days")
    print("=" * 70)

    all_results = []
    checked_athletes = set()  # Track athlete IDs we've already checked
    checked_sports = []  # Track which sports we actually checked

    # Calculate date range for filename
    cutoff_date = datetime.now() - timedelta(days=args.days)
    end_date = datetime.now()

    # Get relevant sports based on current date
    sports_to_check = get_relevant_sports(args.days)

    sport_names = {
        'xc': 'Cross Country',
        'indoor': 'Indoor Track & Field',
        'outdoor': 'Outdoor Track & Field'
    }

    # Apply sport filter if specified
    now = datetime.now()
    any_sport_flag = args.xc or args.indoor or args.outdoor or args.track

    if any_sport_flag:
        # Build list of sports to include based on flags
        selected_sports = set()
        if args.xc:
            selected_sports.add('xc')
        if args.indoor:
            selected_sports.add('indoor')
        if args.outdoor:
            selected_sports.add('outdoor')
        if args.track:
            selected_sports.add('indoor')
            selected_sports.add('outdoor')

        # Filter to selected sports
        sports_to_check = [(s, y) for s, y in sports_to_check if s in selected_sports]

        # If no sports found in season, force them
        if not sports_to_check:
            if 'xc' in selected_sports:
                sports_to_check.append(('xc', now.year))
            if 'indoor' in selected_sports:
                year = now.year if now.month <= 6 else now.year + 1
                sports_to_check.append(('indoor', year))
            if 'outdoor' in selected_sports:
                sports_to_check.append(('outdoor', now.year))

    if not sports_to_check:
        print("No sports to check for the specified criteria.")
        return

    print(f"Checking: {', '.join(f'{sport_names[s]} {y}' for s, y in sports_to_check)}")
    print()

    # Start browser ONCE and reuse it
    options = Options()
    # Use new headless mode - more compatible with modern sites like Angular
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

    # Enable performance logging to capture API tokens
    options.set_capability('goog:loggingPrefs', {'performance': 'ALL'})

    print("Starting browser...")
    print("  Checking ChromeDriver...")
    service = Service(ChromeDriverManager().install())
    print("  Launching Chrome...")
    driver = webdriver.Chrome(service=service, options=options)

    # Initialize API client
    api = AthleticNetAPI()
    api_initialized = False
    use_api = True  # Will be set to False if API fails

    try:
        for sport, year in sports_to_check:
            sport_name = sport_names[sport]
            checked_sports.append(sport)  # Track which sports we checked

            print(f"\n{'='*50}")
            print(f"Checking {sport_name} {year}...")
            print('='*50)

            scraper = AthleticNetScraper(
                headless=args.headless,
                year=year,
                sport=sport,
                days_back=args.days
            )
            # Share the browser instead of starting a new one
            scraper.driver = driver

            # Build team URL for API referer
            team_url = f"https://www.athletic.net/team/65580/{scraper.sport_config['url_path']}/{year}"

            # First, load the team page (needed for both API token capture and Selenium fallback)
            print(f"Loading team page...")
            driver.get(team_url)

            # Wait for page to fully load (important for token capture)
            time.sleep(2)

            # Initialize API on first sport (capture tokens from network logs)
            if not api_initialized and use_api:
                print("Capturing API tokens from browser...")
                if api.init_from_browser(driver):
                    print("  API tokens captured successfully!")
                    api_initialized = True
                else:
                    print("  Could not capture API tokens - will use Selenium scraping")
                    use_api = False

            # ===== NEW: MEET-BASED APPROACH (FASTEST) =====
            # Instead of checking each athlete, we check recent meets directly
            # This is 10-50x faster!

            sport_results = None

            if use_api and api_initialized:
                # Convert sport + year to proper seasonId (indoor uses year + 10000)
                season_id = api.get_season_id(sport, year)

                # For XC: Use FAST meet-based approach (10-50x faster)
                # For Track: Use athlete-based approach (track meet API doesn't return full results)
                if sport == 'xc':
                    print("Using FAST meet-based API approach...")
                    sport_results = api.get_team_results_from_meets(
                        team_id=65580,  # UIS team ID
                        season_id=season_id,
                        sport=sport,
                        cutoff_date=cutoff_date,
                        driver=driver,  # Pass driver for meet page loading
                        referer=team_url
                    )
                else:
                    # Track & Field - use athlete-based approach
                    print("Using athlete-based API approach for track...")
                    roster = api.get_roster(season_id, referer=team_url)
                    if roster:
                        sport_results = api.get_track_results_from_athletes(
                            roster=roster,
                            season_id=season_id,
                            sport=sport,
                            cutoff_date=cutoff_date,
                            driver=driver,
                            referer=team_url
                        )
                    else:
                        sport_results = None

                if sport_results is not None and len(sport_results) > 0:
                    # Success! Add results with sport name
                    for r in sport_results:
                        r['sport'] = sport_name
                    all_results.extend(sport_results)
                    print(f"  Found {len(sport_results)} total results")
                    continue  # Skip to next sport - we're done!
                elif sport_results is not None and len(sport_results) == 0:
                    print("  No results found in the specified time period")
                    continue
                else:
                    print("  API approach failed, falling back to athlete-by-athlete...")

            # ===== FALLBACK: ATHLETE-BY-ATHLETE APPROACH =====
            # Only used if meet-based approach fails

            # Get roster
            roster = None
            if use_api and api_initialized:
                print("Trying API for roster...")
                roster = api.get_roster(year, referer=team_url)
                if roster:
                    print(f"  API: Found {len(roster)} athletes")
                else:
                    print("  API roster failed - falling back to Selenium")

            # Fallback to Selenium scraping for roster
            if not roster:
                roster = scraper.get_roster()

            if not roster:
                print(f"No roster found for {sport_name} {year}")
                continue

            # Filter out already-checked athletes
            new_athletes = [a for a in roster if a['id'] not in checked_athletes]
            skipped = len(roster) - len(new_athletes)

            if skipped > 0:
                print(f"Checking {len(new_athletes)} athletes ({skipped} already checked)...")
            else:
                print(f"Checking {len(new_athletes)} athletes...")

            # Mark all as checked
            for athlete in new_athletes:
                checked_athletes.add(athlete['id'])

            # Process athletes - old API approach (fallback)
            if use_api and api_initialized:
                print(f"Using athlete-by-athlete API approach...")
                api_failed = False

                for i, athlete in enumerate(new_athletes):
                    print(f"  [{i+1}/{len(new_athletes)}] {athlete['name']}...", end=' ', flush=True)

                    referer = f"https://www.athletic.net/athlete/{athlete['id']}/{scraper.sport_config['athlete_path']}"
                    bio_data = api.get_athlete_bio(athlete['id'], sport=sport, referer=referer)

                    if bio_data is None:
                        # API failed - switch to Selenium for remaining athletes
                        print("API failed!")
                        api_failed = True
                        use_api = False
                        remaining = new_athletes[i:]
                        print(f"\nFalling back to Selenium for {len(remaining)} remaining athletes...")
                        break

                    results, bests = api.parse_athlete_results(
                        bio_data, athlete['id'], athlete['name'], cutoff_date, year
                    )

                    if results:
                        print(f"found {len(results)} recent result(s)")
                        for result in results:
                            result['sport'] = sport_name
                            event = result['event']
                            current_time = result['time']

                            # Calculate improvements
                            if event in bests:
                                sr_best = bests[event].get('sr')

                                if result['record_type'] == 'PR' and bests[event].get('pr'):
                                    result['previous_pr'] = bests[event].get('pr')
                                    # Note: API doesn't give us "previous PR", just current PR

                                if result['record_type'] == 'SR' and sr_best:
                                    result['previous_sr'] = sr_best

                                if not result['record_type'] and sr_best:
                                    current_seconds = scraper.time_to_seconds(current_time)
                                    sr_seconds = bests[event].get('sr_seconds', float('inf'))
                                    if sr_seconds != float('inf'):
                                        result['sr_distance'] = (current_seconds - sr_seconds) / sr_seconds * 100
                                        result['current_sr'] = sr_best

                            all_results.append(result)
                    else:
                        print("no recent results")

                if api_failed:
                    # Continue with Selenium for remaining athletes
                    remaining_athletes = new_athletes[new_athletes.index(athlete):]
                else:
                    remaining_athletes = []
            else:
                remaining_athletes = new_athletes

            # Selenium fallback (or primary if API not available)
            if remaining_athletes:
                NUM_PARALLEL_TABS = 3
                for batch_start in range(0, len(remaining_athletes), NUM_PARALLEL_TABS):
                    batch = remaining_athletes[batch_start:batch_start + NUM_PARALLEL_TABS]
                    batch_names = ', '.join(a['name'] for a in batch)
                    print(f"  [{batch_start+1}-{batch_start+len(batch)}/{len(remaining_athletes)}] {batch_names}...", end=' ', flush=True)

                    batch_data = scraper.get_athletes_parallel(batch, num_tabs=NUM_PARALLEL_TABS)

                    found_count = sum(1 for _, results, _ in batch_data if results)
                    if found_count > 0:
                        print(f"found results for {found_count}")
                    else:
                        print("no recent results")

                    for athlete, results, bests in batch_data:
                        if not results:
                            continue

                        for result in results:
                            event = result['event']
                            current_time = result['time']
                            result['sport'] = sport_name

                            # Calculate improvements
                            if event in bests:
                                # For PRs: use previous_pr (second-best all-time) since current PR IS the new time
                                previous_pr = bests[event].get('previous_pr')
                                # For SRs: use previous_sr (second-best this season) since current SR IS the new time
                                previous_sr = bests[event].get('previous_sr')
                                # Current SR for non-PR/SR results
                                sr_best = bests[event].get('sr')

                                if result['record_type'] == 'PR' and previous_pr:
                                    # Validate that previous best is reasonable (similar magnitude to current)
                                    current_secs = scraper.time_to_seconds(current_time)
                                    prev_pr_secs = scraper.time_to_seconds(previous_pr)
                                    # Previous best should be within 50% of current time to be valid
                                    if prev_pr_secs != float('inf') and 0.5 < prev_pr_secs / current_secs < 2.0:
                                        result['pr_improvement'] = scraper.calculate_improvement(current_time, previous_pr)
                                        result['previous_pr'] = previous_pr

                                if result['record_type'] == 'SR' and previous_sr:
                                    # Validate that previous SR is reasonable
                                    current_secs = scraper.time_to_seconds(current_time)
                                    prev_sr_secs = scraper.time_to_seconds(previous_sr)
                                    if prev_sr_secs != float('inf') and 0.5 < prev_sr_secs / current_secs < 2.0:
                                        result['sr_improvement'] = scraper.calculate_improvement(current_time, previous_sr)
                                        result['previous_sr'] = previous_sr

                                if not result['record_type'] and sr_best:
                                    current_seconds = scraper.time_to_seconds(current_time)
                                    sr_seconds = bests[event].get('sr_seconds', float('inf'))
                                    if sr_seconds != float('inf'):
                                        result['sr_distance'] = (current_seconds - sr_seconds) / sr_seconds * 100
                                        result['current_sr'] = sr_best

                            all_results.append(result)

    finally:
        driver.quit()

    if not all_results:
        print("\n" + "=" * 70)
        print("No results found in the specified time period.")
        print("=" * 70)
        return

    # Deduplicate results (same athlete, event, date, time can appear in multiple sports)
    seen = set()
    unique_results = []
    for r in all_results:
        key = (r['athlete_name'], r['event'], r['date_str'], r['time'])
        if key not in seen:
            seen.add(key)
            unique_results.append(r)

    all_results = unique_results
    print(f"After deduplication: {len(all_results)} unique results")

    # Sort results: PRs by improvement, SRs by improvement, others by closeness to SR
    print(f"\nFound {len(all_results)} total results. Sorting...")

    prs = [r for r in all_results if r.get('record_type') == 'PR']
    srs = [r for r in all_results if r.get('record_type') == 'SR']
    fts = [r for r in all_results if r.get('record_type') == 'FT']
    # Separate DNS/DNF from other results
    dns_dnf = [r for r in all_results if not r.get('record_type') and r.get('time', '').upper() in ['DNS', 'DNF']]
    others = [r for r in all_results if not r.get('record_type') and r.get('time', '').upper() not in ['DNS', 'DNF']]

    prs.sort(key=lambda x: x.get('pr_improvement', 0), reverse=True)
    srs.sort(key=lambda x: x.get('sr_improvement', 0), reverse=True)
    fts.sort(key=lambda x: x.get('athlete_name', ''))
    others.sort(key=lambda x: x.get('athlete_name', ''))
    dns_dnf.sort(key=lambda x: x.get('athlete_name', ''))

    # Order: PRs -> SRs -> FTs -> other results -> DNS/DNF
    sorted_results = prs + srs + fts + others + dns_dnf

    # Save to spreadsheet
    data = []
    for r in sorted_results:
        prev_pr = r.get('previous_pr')
        prev_sr = r.get('previous_sr')
        pr_improvement = r.get('pr_improvement', 0)
        sr_improvement = r.get('sr_improvement', 0)

        row = {
            'Name': r['athlete_name'],
            'Type': r.get('record_type') or '-',
            'Sport': r.get('sport', ''),
            'Event': r['event'],
            'Time/Mark': r['time'],
            'Place': r['place'],
            'Date': r['date_str'],
            'Meet': r['meet_name'],
        }

        # Previous Best (PR) column
        if r.get('record_type') == 'FT' or r.get('first_at_distance'):
            row['Previous Best'] = '-'
        elif prev_pr:
            row['Previous Best'] = prev_pr
        else:
            row['Previous Best'] = '-'

        # Previous Season Best column
        if prev_sr:
            row['Previous SR'] = prev_sr
        else:
            row['Previous SR'] = '-'

        # % Improvement from PR column
        if r.get('record_type') == 'PR' and prev_pr:
            row['% from PR'] = f"{pr_improvement:.2f}%"
        elif r.get('record_type') == 'FT' or not prev_pr:
            row['% from PR'] = '-'
        elif pr_improvement > 0:
            row['% from PR'] = f"{pr_improvement:.2f}%"
        elif pr_improvement < 0:
            row['% from PR'] = f"{pr_improvement:.2f}%"
        else:
            row['% from PR'] = '-'

        # % Improvement from SR column
        if r.get('record_type') == 'SR' and prev_sr:
            row['% from SR'] = f"{sr_improvement:.2f}%"
        elif not prev_sr:
            row['% from SR'] = '-'
        elif sr_improvement != 0:
            row['% from SR'] = f"{sr_improvement:.2f}%"
        else:
            row['% from SR'] = '-'

        # NCAA D2 Standard columns
        ncaa_standard = r.get('ncaa_standard')
        ncaa_diff = r.get('ncaa_diff')
        ncaa_diff_pct = r.get('ncaa_diff_pct')

        if ncaa_standard:
            # Format the standard as a readable time/distance
            is_field = any(f in r['event'].lower() for f in ['jump', 'vault', 'put', 'throw', 'discus', 'hammer', 'javelin'])
            if is_field:
                row['NCAA Std'] = f"{ncaa_standard:.2f}m"
            else:
                row['NCAA Std'] = format_standard_time(ncaa_standard)

            # Format the difference as percentage
            # Negative = qualified (under standard for time, over for field)
            if ncaa_diff_pct is not None:
                if is_field:
                    # Field: positive % = over standard (good/qualified)
                    # Flip sign so negative = not qualified, positive = qualified
                    display_pct = ncaa_diff_pct
                    if display_pct >= 0:
                        row['vs NCAA'] = f"+{display_pct:.1f}%"
                    else:
                        row['vs NCAA'] = f"{display_pct:.1f}%"
                else:
                    # Time: negative % = under standard (qualified)
                    if ncaa_diff_pct <= 0:
                        row['vs NCAA'] = f"{ncaa_diff_pct:.1f}%"
                    else:
                        row['vs NCAA'] = f"+{ncaa_diff_pct:.1f}%"
            else:
                row['vs NCAA'] = '-'
        else:
            row['NCAA Std'] = '-'
            row['vs NCAA'] = '-'

        data.append(row)

    df = pd.DataFrame(data)
    columns = ['Name', 'Type', 'Sport', 'Event', 'Time/Mark', 'Place', 'Date', 'Meet', 'Previous Best', 'Previous SR', '% from PR', '% from SR', 'NCAA Std', 'vs NCAA']
    df = df[columns]

    # Build filename with sport(s) and date range
    sport_abbrevs = {'xc': 'XC', 'indoor': 'Indoor', 'outdoor': 'Outdoor'}
    sports_str = '_'.join(sport_abbrevs.get(s, s) for s in dict.fromkeys(checked_sports))  # Preserve order, remove dupes
    start_str = cutoff_date.strftime('%b%d')
    end_str = end_date.strftime('%b%d')
    base_filename = f"results_{sports_str}_{start_str}-{end_str}"
    filepath = f"{output_dir}/{base_filename}.xlsx"

    # Try to save, overwriting any existing file
    # Export to JSON first (important for cloud mode)
    try:
        _push_results_to_website(data, cutoff_date, end_date, checked_sports, cloud_mode=args.cloud)
    except Exception as e:
        print(f"\nWarning: Could not push to website: {e}")

    # Try to save Excel file
    import os
    try:
        # If file exists, check if it's writable
        if os.path.exists(filepath):
            with open(filepath, 'a'):
                pass  # Just checking if file is locked
        _save_styled_excel(df, filepath, sorted_results)
        print(f"\nResults saved to: {filepath}")
    except (PermissionError, OSError) as e:
        if not args.cloud:
            print(f"\nError: Could not save to {filepath}")
            print(f"The file may be open in another application (like Excel).")
            print(f"Please close the file and run the scraper again.")
            return
        else:
            print(f"\nSkipping Excel save in cloud mode")

    print(f"  PRs: {len(prs)}")
    print(f"  SRs: {len(srs)}")
    print(f"  First Times: {len(fts)}")
    print(f"  Other Results: {len(others)}")
    print(f"  DNS/DNF: {len(dns_dnf)}")

    print("\n" + "=" * 70)
    print("SUCCESS! Check the spreadsheet for results.")
    print("=" * 70)


if __name__ == "__main__":
    main()
